#!/usr/bin/env python3
"""Validate the structure and internal consistency of the research artifact."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from collections import Counter
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook
from PIL import Image
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
RELEASE_VERSION = "1.0.0"
RELEASE_DATE = "2026-08-06"
VALIDATION_DATE = "2026-08-07"
CONTROLLED_LABELS = {
    "Abbreviation",
    "Acronym",
    "Affixation",
    "Blending",
    "Borrowing",
    "Clipping",
    "Code Word",
    "Community-specific Jargon",
    "Compound",
    "Functional Shift",
    "Initialism",
    "Meme Expression",
    "Metaphor",
    "Multiword Expression",
    "Semantic Shift",
}
JSON_FIELDS = [
    "id",
    "domain",
    "term",
    "linguistic_construction",
    "community_subcommunity",
    "real_usage_example",
    "source_url",
]
WORKBOOK_HEADERS = [
    "Term",
    "Linguistic Construction",
    "Community / Sub-community",
    "Real Usage Example",
    "Source Link",
]
REQUIRED_PATHS = [
    "README.md",
    "CITATION.cff",
    "CHANGELOG.md",
    "LICENSE_NOTES.md",
    ".gitignore",
    "requirements.txt",
    "data/source/refined_terminology_usage.xlsx",
    "data/release/community_terminology_dataset.json",
    "documentation/annotation_guideline.docx",
    "documentation/annotation_guideline.pdf",
    "documentation/pura_final_report.docx",
    "documentation/pura_final_report.pdf",
    "docs/dataset_card.md",
    "docs/methodology.md",
    "docs/annotation_schema.md",
    "docs/limitations_and_ethics.md",
    "docs/future_experiments.md",
    "docs/file_manifest.md",
    "assets/domain_distribution.png",
    "assets/linguistic_construction_distribution.png",
    "scripts/validate_artifact.py",
    "scripts/generate_summary.py",
    "validation/validation_report.md",
    "validation/dataset_summary.json",
    "validation/checksums.sha256",
    "archive/original_files/Community_Terminology_Annotation_Guideline(1).docx",
    "archive/original_files/Community_Terminology_Annotation_Guideline.pdf",
    "archive/original_files/community_terminology_dataset(1).json",
    "archive/original_files/PURA Final Report Guanjun Yan.docx",
    "archive/original_files/Refined_Terminology_Usage.xlsx",
    ".github/workflows/validate.yml",
]
PUBLICATION_ARCHIVE_PAIRS = [
    ("documentation/annotation_guideline.docx", "archive/original_files/Community_Terminology_Annotation_Guideline(1).docx"),
    ("documentation/annotation_guideline.pdf", "archive/original_files/Community_Terminology_Annotation_Guideline.pdf"),
    ("documentation/pura_final_report.docx", "archive/original_files/PURA Final Report Guanjun Yan.docx"),
    ("data/source/refined_terminology_usage.xlsx", "archive/original_files/Refined_Terminology_Usage.xlsx"),
]
README_REQUIRED_LINKS = {
    "data/source/refined_terminology_usage.xlsx",
    "data/release/community_terminology_dataset.json",
    "documentation/annotation_guideline.pdf",
    "documentation/annotation_guideline.docx",
    "documentation/pura_final_report.pdf",
    "documentation/pura_final_report.docx",
    "docs/dataset_card.md",
    "docs/methodology.md",
    "docs/annotation_schema.md",
    "docs/limitations_and_ethics.md",
    "docs/future_experiments.md",
    "docs/file_manifest.md",
    "validation/validation_report.md",
    "assets/domain_distribution.png",
    "assets/linguistic_construction_distribution.png",
    "CITATION.cff",
    "LICENSE_NOTES.md",
}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


class Validation:
    def __init__(self) -> None:
        self.passes: list[str] = []
        self.warnings: list[str] = []
        self.failures: list[str] = []

    def check(self, condition: bool, success: str, failure: str) -> None:
        (self.passes if condition else self.failures).append(success if condition else failure)


def validate_files(result: Validation) -> None:
    missing = [path for path in REQUIRED_PATHS if not (ROOT / path).is_file()]
    result.check(not missing, "All required repository files are present.", f"Missing required files: {missing}")
    for publication, archived in PUBLICATION_ARCHIVE_PAIRS:
        if (ROOT / publication).is_file() and (ROOT / archived).is_file():
            result.check(
                sha256(ROOT / publication) == sha256(ROOT / archived),
                f"Publication copy is byte-identical to its archived source: {publication}",
                f"Publication copy differs from archived source: {publication}",
            )


def load_workbook_records(result: Validation) -> tuple[list[dict], dict[str, int]]:
    workbook = load_workbook(ROOT / "data/source/refined_terminology_usage.xlsx", read_only=True, data_only=False)
    result.check(
        workbook.sheetnames == ["Cybersecurity", "Gaming", "Source Audit"],
        "Workbook contains the expected sheets in the expected order.",
        f"Unexpected workbook sheets: {workbook.sheetnames}",
    )
    records: list[dict] = []
    counts: dict[str, int] = {}
    for sheet_name, domain in (("Cybersecurity", "cybersecurity"), ("Gaming", "gaming")):
        sheet = workbook[sheet_name]
        headers = [sheet.cell(1, column).value for column in range(1, 6)]
        result.check(
            headers == WORKBOOK_HEADERS,
            f"{sheet_name} has the expected five headers.",
            f"{sheet_name} headers differ: {headers}",
        )
        rows = []
        for values in sheet.iter_rows(min_row=2, max_col=5, values_only=True):
            if all(blank(value) for value in values):
                continue
            rows.append(
                {
                    "domain": domain,
                    "term": values[0],
                    "linguistic_construction": values[1],
                    "community_subcommunity": values[2],
                    "real_usage_example": values[3],
                    "source_url": values[4],
                }
            )
        counts[domain] = len(rows)
        records.extend(rows)
    return records, counts


def validate_data(result: Validation) -> dict:
    payload = json.loads((ROOT / "data/release/community_terminology_dataset.json").read_text(encoding="utf-8"))
    archived_payload = json.loads(
        (ROOT / "archive/original_files/community_terminology_dataset(1).json").read_text(encoding="utf-8")
    )
    normalized_archive = dict(archived_payload)
    normalized_archive["version"] = RELEASE_VERSION
    result.check(
        payload == normalized_archive,
        "Publication JSON differs from the archived original only by the documented version normalization.",
        "Publication JSON contains changes beyond the documented version normalization.",
    )
    result.check(
        isinstance(payload, dict) and isinstance(payload.get("records"), list),
        "JSON release is a metadata object containing a records array.",
        "JSON release is not a metadata object containing a records array.",
    )
    records = payload.get("records", [])
    result.check(len(records) == 181, "JSON contains 181 records.", f"JSON contains {len(records)} records, expected 181.")
    domain_counts = Counter(record.get("domain") for record in records)
    result.check(
        domain_counts == Counter({"cybersecurity": 74, "gaming": 107}),
        "JSON domain counts are cybersecurity=74 and gaming=107.",
        f"Unexpected JSON domain counts: {dict(domain_counts)}",
    )
    missing = {
        field: sum(field not in record or blank(record.get(field)) for record in records) for field in JSON_FIELDS
    }
    result.check(not any(missing.values()), "All required JSON fields are present and non-empty.", f"Missing values: {missing}")
    field_sets_ok = all(set(record) == set(JSON_FIELDS) for record in records)
    result.check(field_sets_ok, "Every JSON record has exactly the documented seven fields.", "JSON record fields differ from the schema.")
    ids = [record.get("id") for record in records]
    result.check(len(ids) == len(set(ids)), "JSON record IDs are unique.", "Duplicate JSON record IDs were found.")
    expected_ids = [f"cybersecurity_{i:03d}" for i in range(1, 75)] + [f"gaming_{i:03d}" for i in range(1, 108)]
    result.check(ids == expected_ids, "JSON IDs are complete and ordered by domain.", "JSON IDs are missing, malformed, or out of order.")
    observed_labels = {record.get("linguistic_construction") for record in records}
    result.check(
        observed_labels == CONTROLLED_LABELS,
        "Observed labels exactly match the 15-label controlled vocabulary.",
        f"Label mismatch: unexpected={sorted(observed_labels - CONTROLLED_LABELS)}, missing={sorted(CONTROLLED_LABELS - observed_labels)}",
    )
    invalid_urls = [record.get("id") for record in records if urlparse(str(record.get("source_url"))).scheme not in {"http", "https"}]
    result.check(not invalid_urls, "All source URLs use HTTP or HTTPS syntax.", f"Invalid source URL syntax in IDs: {invalid_urls}")

    workbook_records, workbook_counts = load_workbook_records(result)
    result.check(
        workbook_counts == {"cybersecurity": 74, "gaming": 107},
        "Workbook domain counts are cybersecurity=74 and gaming=107.",
        f"Unexpected workbook domain counts: {workbook_counts}",
    )
    mismatches = []
    if len(workbook_records) != len(records):
        mismatches.append("record_count")
    else:
        for index, (workbook_record, json_record) in enumerate(zip(workbook_records, records), start=1):
            for field in workbook_record:
                if workbook_record[field] != json_record.get(field):
                    mismatches.append(f"record {index} field {field}")
    result.check(not mismatches, "Workbook and JSON agree exactly across all 181 records.", f"Cross-format mismatches: {mismatches[:10]}")

    terms = Counter(record["term"] for record in records)
    urls = Counter(record["source_url"] for record in records)
    duplicate_terms = {term: count for term, count in terms.items() if count > 1}
    duplicate_urls = {url: count for url, count in urls.items() if count > 1}
    result.check(not duplicate_terms, "No exact duplicate terms were found.", f"Exact duplicate terms: {duplicate_terms}")
    if duplicate_urls == {
        "https://www.reddit.com/r/it/comments/1pmohm0/mfa_fatigue_attacks_are_getting_out_of_control/": 2
    }:
        result.warnings.append(
            "One source URL is shared by two distinct records (MFA fatigue and MFA bombing); this is documented and retained."
        )
    else:
        result.warnings.append(f"Duplicate source URL groups differ from the documented release: {duplicate_urls}")

    metadata = payload.get("statistics", {})
    result.check(metadata.get("total_records") == len(records), "Embedded total statistic matches recomputation.", "Embedded total statistic is stale.")
    result.check(metadata.get("records_by_domain") == dict(domain_counts), "Embedded domain statistics match recomputation.", "Embedded domain statistics are stale.")
    construction_counts = Counter(record["linguistic_construction"] for record in records)
    result.check(
        metadata.get("records_by_linguistic_construction") == dict(construction_counts),
        "Embedded construction statistics match recomputation.",
        "Embedded construction statistics are stale.",
    )
    return {"records": records, "domain_counts": dict(domain_counts), "construction_counts": dict(construction_counts)}


def validate_documents_and_assets(result: Validation) -> None:
    for relative_path, expected_pages in (
        ("documentation/annotation_guideline.pdf", 2),
        ("documentation/pura_final_report.pdf", 3),
    ):
        path = ROOT / relative_path
        if not path.exists():
            continue
        reader = PdfReader(path)
        result.check(len(reader.pages) == expected_pages, f"{relative_path} has {expected_pages} pages.", f"{relative_path} has {len(reader.pages)} pages, expected {expected_pages}.")
        blank_pages = [i for i, page in enumerate(reader.pages, start=1) if not (page.extract_text() or "").strip()]
        result.check(not blank_pages, f"{relative_path} has no text-empty pages.", f"{relative_path} has text-empty pages: {blank_pages}")
    for relative_path in ("assets/domain_distribution.png", "assets/linguistic_construction_distribution.png"):
        path = ROOT / relative_path
        if not path.exists():
            continue
        with Image.open(path) as image:
            width, height = image.size
            result.check(width >= 1000 and height >= 600, f"{relative_path} has publication-readable dimensions ({width}x{height}).", f"{relative_path} is too small ({width}x{height}).")


def validate_readme(result: Validation) -> None:
    readme_path = ROOT / "README.md"
    if not readme_path.exists():
        return
    readme = readme_path.read_text(encoding="utf-8")
    word_count = len(re.findall(r"\b[\w'-]+\b", re.sub(r"```.*?```", "", readme, flags=re.S)))
    result.check(1500 <= word_count <= 2500, f"README prose length is within target ({word_count} words).", f"README prose length is outside 1,500-2,500 words ({word_count}).")
    headings = [
        "Project overview",
        "Research question",
        "Why community-specific terminology matters",
        "Artifact contents",
        "Dataset at a glance",
        "Dataset schema",
        "Collection and annotation workflow",
        "Linguistic-construction taxonomy",
        "Repository structure",
        "Using the dataset",
        "Validating the release",
        "Completed project outcomes",
        "Proposed future experiments",
        "Limitations and ethical considerations",
        "Citation",
        "Acknowledgements",
    ]
    missing = [heading for heading in headings if not re.search(rf"^##+\s+{re.escape(heading)}\s*$", readme, re.M | re.I)]
    result.check(not missing, "README contains all required sections.", f"README is missing required sections: {missing}")
    targets = set(re.findall(r"!?\[[^\]]*\]\(([^)#]+)(?:#[^)]+)?\)", readme))
    missing_links = sorted(README_REQUIRED_LINKS - targets)
    broken_links = sorted(
        target
        for target in targets
        if not urlparse(target).scheme and not target.startswith("#") and not (ROOT / target).exists()
    )
    result.check(
        not missing_links and not broken_links,
        "README contains all required artifact links and every relative target exists.",
        f"README link issues: missing={missing_links}, broken={broken_links}",
    )


def validate_release_metadata(result: Validation) -> None:
    checks = {
        "data/release/community_terminology_dataset.json": (
            json.loads((ROOT / "data/release/community_terminology_dataset.json").read_text(encoding="utf-8")).get("version")
            == RELEASE_VERSION
        ),
        "validation/dataset_summary.json": (
            json.loads((ROOT / "validation/dataset_summary.json").read_text(encoding="utf-8")).get("release_version")
            == RELEASE_VERSION
        ),
        "CITATION.cff": (
            f"version: {RELEASE_VERSION}" in (ROOT / "CITATION.cff").read_text(encoding="utf-8")
            and f"date-released: {RELEASE_DATE}" in (ROOT / "CITATION.cff").read_text(encoding="utf-8")
        ),
        "CHANGELOG.md": (
            f"## [{RELEASE_VERSION}] - {RELEASE_DATE}" in (ROOT / "CHANGELOG.md").read_text(encoding="utf-8")
        ),
        "README.md": (
            f"Version {RELEASE_VERSION}" in (ROOT / "README.md").read_text(encoding="utf-8")
        ),
    }
    inconsistent = sorted(path for path, valid in checks.items() if not valid)
    result.check(
        not inconsistent,
        f"Release metadata is consistent at version {RELEASE_VERSION} dated {RELEASE_DATE}.",
        f"Inconsistent release metadata: {inconsistent}",
    )


def validate_checksums(result: Validation) -> None:
    path = ROOT / "validation/checksums.sha256"
    if not path.exists():
        return
    bad = []
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.strip() or line.startswith("#"):
            continue
        expected, relative_path = line.split(maxsplit=1)
        relative_path = relative_path.lstrip("*")
        target = ROOT / relative_path
        if not target.is_file() or sha256(target) != expected:
            bad.append(relative_path)
    result.check(not bad, "All recorded SHA-256 checksums match.", f"Checksum mismatches: {bad}")


def write_report(result: Validation) -> None:
    lines = [
        "# Validation report",
        "",
        f"Release: **{RELEASE_VERSION}**",
        "",
        f"Release date: **{RELEASE_DATE}**",
        "",
        f"Validation date: **{VALIDATION_DATE}**",
        "",
        "## Outcome",
        "",
        f"**{'PASS' if not result.failures else 'FAIL'}** - {len(result.passes)} checks passed, {len(result.warnings)} warning(s), and {len(result.failures)} failure(s).",
        "",
        "## Passed checks",
        "",
    ]
    lines.extend(f"- {message}" for message in result.passes)
    lines.extend(["", "## Warnings", ""])
    lines.extend(f"- {message}" for message in result.warnings or ["None."])
    lines.extend(["", "## Failures", ""])
    lines.extend(f"- {message}" for message in result.failures or ["None."])
    lines.extend(
        [
            "",
            "## Scope and interpretation",
            "",
            "This validation checks structure, counts, controlled labels, required fields, cross-format equality, file identity or documented record-level equivalence, release metadata, README links, PDF page/text presence, image dimensions, and recorded checksums. It validates URL syntax but deliberately does not require live URL retrieval: community pages can be edited, deleted, blocked, or rate-limited. It also does not treat automated checks as a substitute for additional expert annotation or inter-annotator agreement measurement.",
            "",
        ]
    )
    destination = ROOT / "validation/validation_report.md"
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--write-report", action="store_true", help="Refresh validation/validation_report.md.")
    args = parser.parse_args()
    result = Validation()
    validate_files(result)
    validate_data(result)
    validate_documents_and_assets(result)
    validate_readme(result)
    validate_release_metadata(result)
    validate_checksums(result)
    if args.write_report:
        write_report(result)
    for message in result.passes:
        print(f"PASS: {message}")
    for message in result.warnings:
        print(f"WARN: {message}")
    for message in result.failures:
        print(f"FAIL: {message}")
    print(f"\nSummary: {len(result.passes)} passed, {len(result.warnings)} warning(s), {len(result.failures)} failure(s).")
    return 1 if result.failures else 0


if __name__ == "__main__":
    sys.exit(main())
