#!/usr/bin/env python3
"""Generate deterministic release statistics and publication charts."""

from __future__ import annotations

import argparse
import json
from collections import Counter, defaultdict
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATASET_PATH = ROOT / "data" / "release" / "community_terminology_dataset.json"
WORKBOOK_PATH = ROOT / "data" / "source" / "refined_terminology_usage.xlsx"
SUMMARY_PATH = ROOT / "validation" / "dataset_summary.json"
ASSET_DIR = ROOT / "assets"


def load_records() -> tuple[dict, list[dict]]:
    payload = json.loads(DATASET_PATH.read_text(encoding="utf-8"))
    if not isinstance(payload, dict) or not isinstance(payload.get("records"), list):
        raise ValueError("The JSON release must be an object containing a records array.")
    return payload, payload["records"]


def duplicate_groups(records: list[dict], field: str) -> dict[str, list[str]]:
    groups: dict[str, list[str]] = defaultdict(list)
    for record in records:
        groups[record[field]].append(record["id"])
    return {value: ids for value, ids in groups.items() if len(ids) > 1}


def source_family_counts() -> dict[str, dict[str, int]]:
    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    sheet = workbook["Source Audit"]
    cybersecurity: dict[str, int] = {}
    gaming: dict[str, int] = {}
    for row in range(9, 17):
        cyber_family = sheet.cell(row=row, column=1).value
        cyber_count = sheet.cell(row=row, column=2).value
        game_family = sheet.cell(row=row, column=4).value
        game_count = sheet.cell(row=row, column=5).value
        if cyber_family:
            cybersecurity[str(cyber_family)] = int(cyber_count)
        if game_family:
            gaming[str(game_family)] = int(game_count)
    return {"cybersecurity": cybersecurity, "gaming": gaming}


def build_summary(records: list[dict]) -> dict:
    domain_counts = Counter(record["domain"] for record in records)
    construction_counts = Counter(record["linguistic_construction"] for record in records)
    host_counts = Counter(
        urlparse(record["source_url"]).netloc.lower().removeprefix("www.") for record in records
    )
    fields = [
        "id",
        "domain",
        "term",
        "linguistic_construction",
        "community_subcommunity",
        "real_usage_example",
        "source_url",
    ]
    missing = {
        field: sum(record.get(field) is None or str(record.get(field)).strip() == "" for record in records)
        for field in fields
    }
    duplicate_terms = duplicate_groups(records, "term")
    duplicate_urls = duplicate_groups(records, "source_url")
    return {
        "dataset_name": "Community-Specific Gaming and Cybersecurity Terminology",
        "release_version": "1.0.0",
        "total_records": len(records),
        "records_by_domain": dict(sorted(domain_counts.items())),
        "records_by_linguistic_construction": dict(construction_counts.most_common()),
        "records_by_source_hostname": dict(host_counts.most_common()),
        "records_by_source_family": source_family_counts(),
        "missing_values_by_field": missing,
        "duplicate_terms": {
            "group_count": len(duplicate_terms),
            "record_count": sum(len(ids) for ids in duplicate_terms.values()),
            "groups": duplicate_terms,
        },
        "duplicate_source_urls": {
            "group_count": len(duplicate_urls),
            "record_count": sum(len(ids) for ids in duplicate_urls.values()),
            "groups": duplicate_urls,
        },
    }


def make_charts(summary: dict) -> None:
    from PIL import Image, ImageDraw, ImageFont

    ASSET_DIR.mkdir(parents=True, exist_ok=True)

    def font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont:
        names = (
            [
                "/System/Library/Fonts/Supplemental/Arial Bold.ttf",
                "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
                "C:/Windows/Fonts/arialbd.ttf",
            ]
            if bold
            else [
                "/System/Library/Fonts/Supplemental/Arial.ttf",
                "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
                "C:/Windows/Fonts/arial.ttf",
            ]
        )
        for filename in names:
            if Path(filename).is_file():
                return ImageFont.truetype(filename, size=size)
        return ImageFont.load_default(size=size)

    background = "#FFFFFF"
    ink = "#253247"
    grid = "#D8DEE8"
    blue = "#2A6F97"
    orange = "#E69F00"

    domains = ["Cybersecurity", "Gaming"]
    values = [
        summary["records_by_domain"]["cybersecurity"],
        summary["records_by_domain"]["gaming"],
    ]
    image = Image.new("RGB", (1600, 900), background)
    draw = ImageDraw.Draw(image)
    draw.text((800, 70), "Community terminology records by domain", fill=ink, font=font(46, True), anchor="mm")
    left, top, right, bottom = 210, 170, 1480, 730
    max_value = 120
    for tick in range(0, max_value + 1, 20):
        y = bottom - (tick / max_value) * (bottom - top)
        draw.line((left, y, right, y), fill=grid, width=2)
        draw.text((left - 25, y), str(tick), fill=ink, font=font(24), anchor="rm")
    draw.line((left, top, left, bottom), fill=ink, width=3)
    draw.line((left, bottom, right, bottom), fill=ink, width=3)
    bar_width = 330
    centers = [600, 1110]
    for label, value, color, center in zip(domains, values, (blue, orange), centers):
        bar_height = (value / max_value) * (bottom - top)
        draw.rounded_rectangle(
            (center - bar_width / 2, bottom - bar_height, center + bar_width / 2, bottom),
            radius=8,
            fill=color,
        )
        draw.text((center, bottom - bar_height - 24), str(value), fill=ink, font=font(32, True), anchor="ms")
        draw.text((center, bottom + 45), label, fill=ink, font=font(28), anchor="ma")
    axis_label = Image.new("RGBA", (360, 60), (255, 255, 255, 0))
    axis_draw = ImageDraw.Draw(axis_label)
    axis_draw.text((180, 30), "Number of records", fill=ink, font=font(28), anchor="mm")
    axis_label = axis_label.rotate(90, expand=True)
    image.paste(axis_label, (35, (900 - axis_label.height) // 2), axis_label)
    image.save(ASSET_DIR / "domain_distribution.png", optimize=True)

    construction_counts = summary["records_by_linguistic_construction"]
    ordered = sorted(construction_counts.items(), key=lambda item: (item[1], item[0]))
    labels = [label for label, _ in ordered]
    counts = [count for _, count in ordered]
    image = Image.new("RGB", (1900, 1350), background)
    draw = ImageDraw.Draw(image)
    draw.text(
        (950, 62),
        "Community terminology records by linguistic construction",
        fill=ink,
        font=font(44, True),
        anchor="mm",
    )
    left, top, right, bottom = 540, 145, 1760, 1210
    max_value = 80
    for tick in range(0, max_value + 1, 10):
        x = left + (tick / max_value) * (right - left)
        draw.line((x, top, x, bottom), fill=grid, width=2)
        draw.text((x, bottom + 24), str(tick), fill=ink, font=font(20), anchor="ma")
    row_height = (bottom - top) / len(labels)
    for index, (label, count) in enumerate(zip(labels, counts)):
        center_y = top + row_height * (index + 0.5)
        bar_height = row_height * 0.62
        bar_width = (count / max_value) * (right - left)
        draw.text((left - 22, center_y), label, fill=ink, font=font(22), anchor="rm")
        draw.rounded_rectangle(
            (left, center_y - bar_height / 2, left + bar_width, center_y + bar_height / 2),
            radius=5,
            fill=blue,
        )
        draw.text((left + bar_width + 14, center_y), str(count), fill=ink, font=font(21, True), anchor="lm")
    draw.line((left, top, left, bottom), fill=ink, width=3)
    draw.line((left, bottom, right, bottom), fill=ink, width=3)
    draw.text(((left + right) / 2, 1300), "Number of records", fill=ink, font=font(27), anchor="mm")
    image.save(ASSET_DIR / "linguistic_construction_distribution.png", optimize=True)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--check",
        action="store_true",
        help="Fail if the committed summary differs from a fresh computation.",
    )
    args = parser.parse_args()
    _, records = load_records()
    summary = build_summary(records)
    rendered = json.dumps(summary, indent=2, ensure_ascii=False) + "\n"
    if args.check:
        if not SUMMARY_PATH.exists() or SUMMARY_PATH.read_text(encoding="utf-8") != rendered:
            print("ERROR: validation/dataset_summary.json is stale. Run scripts/generate_summary.py.")
            return 1
        print("PASS: dataset_summary.json matches the release data.")
        return 0

    SUMMARY_PATH.parent.mkdir(parents=True, exist_ok=True)
    SUMMARY_PATH.write_text(rendered, encoding="utf-8")
    make_charts(summary)
    print(f"Wrote {SUMMARY_PATH.relative_to(ROOT)} and two charts in assets/.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
